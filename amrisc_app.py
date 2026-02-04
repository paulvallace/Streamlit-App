
# Amrisc_SOV_app.py
# Streamlit UI wrapper for AmRisc SOV processing, modeled after your CC_app
# - Detects source sheet/header (Street/City/State/Zip synonyms)
# - Maps source columns to AmRisc SOV-APP targets
# - Finds the real template header row (top-100 scan); writes below it
# - Safe writing around merged cells; optional append to first empty row
# - Strict sprinkler mapping: Y/N -> "Sprinklered (Y/N)", % -> "Percent Sprinklered"
# - Download completed workbook + optional save to disk

import io
import re
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# =========================
# Page Setup
# =========================
st.set_page_config(page_title="AmRisc SOV Builder", page_icon="📄", layout="wide")
st.title("📄 AmRisc SOV Builder")
st.caption("Point-and-click app that normalizes a client SOV into the AmRisc SOV-APP template.")

# =========================
# Helpers (normalized & merged-cell-safe)
# =========================
def normalize_alias(x: Optional[str]) -> str:
    """Lower, trim, collapse whitespace, '&amp;'->'and', remove *,(), strip non-alphanum."""
    if x is None:
        return ""
    s = str(x).strip().lower().replace("&amp;", "and")
    s = re.sub(r"[\*\(\)]", "", s)      # remove *, (, )
    s = re.sub(r"\s+", " ", s)         # collapse whitespace
    return re.sub(r"[^a-z0-9]", "", s)  # strip non-alphanum

def split_lines_safe(s: Optional[str]):
    """Split on CR/LF and trim parts."""
    if not isinstance(s, str):
        return []
    parts = re.split(r"[\r\n]+", s)
    return [p.strip() for p in parts if p and p.strip()]

def is_blank_series(series: pd.Series) -> pd.Series:
    return series.isna() | (series.astype(str).str.strip() == "")

def split_city_state_zip_col(series: pd.Series) -> pd.DataFrame:
    """Split 'City, ST, 12345[-6789]' -> City/State/Zip (all string dtype)."""
    s = series.astype("string").str.strip()
    pat = r"^\s*(?P<City>[^,]+?)\s*,\s*(?P<State>[A-Za-z]{2})\s*,\s*(?P<Zip>\d{5}(?:-\d{4})?)\s*$"
    parts = s.str.extract(pat)
    for col in ("City", "State", "Zip"):
        if col in parts:
            parts[col] = parts[col].astype("string").str.strip()
    if "State" in parts:
        parts["State"] = parts["State"].str.upper()
    return parts

def looks_like_header(row_vals, min_groups=3) -> bool:
    """Detect header row by presence of Street/City/State/Zip tokens."""
    toks = {normalize_alias(v) for v in row_vals if isinstance(v, (str, int, float)) and str(v).strip()}
    groups = [
        {"street", "streetaddress", "address", "Address 1", "location address"},
        {"city", "town"},
        {"state", "statecode", "province"},
        {"zip", "zipcode", "postal", "postalcode"},
    ]
    score = sum(any(alias in toks for alias in g) for g in groups)
    return score >= min_groups

def find_sheet_and_header(xlsx_file_like, search_rows=50) -> Tuple[str, int]:
    """
    Return (sheet_name, header_row_index) by scanning each sheet for likely
    address headers (Street/City/State/Zip).
    """
    data = xlsx_file_like.getvalue() if isinstance(xlsx_file_like, io.BytesIO) else xlsx_file_like.read()
    buf1 = io.BytesIO(data)
    xfile = pd.ExcelFile(buf1, engine="openpyxl")
    for sheet in xfile.sheet_names:
        buf2 = io.BytesIO(data)  # fresh handle for each read
        raw = pd.read_excel(buf2, sheet_name=sheet, header=None, engine="openpyxl")
        for i in range(min(search_rows, len(raw))):
            if looks_like_header(raw.iloc[i].tolist()):
                return sheet, i
    raise RuntimeError("Could not find a header row in any sheet (looking for Street/City/State/Zip synonyms).")

def build_alias_to_colidx(ws, header_row: int) -> tuple[dict, list]:
    """Map normalized header aliases to 1-based column indices (handles wrapped headers)."""
    alias_to_colidx = {}
    raw_headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    for idx, header in enumerate(raw_headers, start=1):
        if header is None:
            continue
        aliases = set()
        aliases.add(header)
        if isinstance(header, str):
            aliases.add(header.replace("\n", " ").replace("\r", " "))
            for part in split_lines_safe(header):
                aliases.add(part)
        more = set()
        for a in aliases:
            if isinstance(a, str):
                more.add(a.replace("&", "and"))
        aliases |= more
        for a in aliases:
            key = normalize_alias(a)
            if key and key not in alias_to_colidx:
                alias_to_colidx[key] = idx
    return alias_to_colidx, raw_headers

def detect_template_header_row(ws, targets_norm: set, scan_top=100) -> int:
    """Find the row with max matches against target labels."""
    best_row, best_score = None, -1
    for r in range(1, scan_top + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        toks = {normalize_alias(v) for v in row_vals if v}
        score = sum(k in toks for k in targets_norm)
        if score > best_score:
            best_row, best_score = r, score
    return best_row or 1

def first_empty_row_under(ws, column_index: int, start: int = 3) -> int:
    r = start
    while True:
        if ws.cell(row=r, column=column_index).value in (None, ""):
            return r
        r += 1

def safe_write(ws, row: int, col: int, value):
    """Write value; if target is merged, write into the merged range's top-left cell."""
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col, value=value)
                break
    else:
        c.value = value

def non_empty(s: Optional[str]) -> bool:
    return isinstance(s, str) and s.strip() != ""

def find_sprinkler_col(df: pd.DataFrame) -> Optional[str]:
    """Best-effort: find a sprinkler column by name."""
    for c in df.columns:
        cl = str(c).strip().lower()
        if any(k in cl for k in ["sprink", "sprinklered", "sprinkler", "sprnkl"]):
            return c
    return None

# --- Strict sprinkler mapping (map what's present; no derivations) ---
def map_sprinkler_to_targets(
    df: pd.DataFrame,
    source_col_name: str,
    yn_target_col: str = "Sprinklered (Y/N)",
    pct_target_col: str = "Percent Sprinklered"
) -> pd.DataFrame:
    """
    If values look like Y/N, fill 'Sprinklered (Y/N)' with 'Y'/'N'.
    If values look like percentages (0..100 or with %), fill 'Percent Sprinklered' with 0..100.
    No derivations between Y/N and %.
    """
    # Ensure targets exist
    if yn_target_col not in df.columns:
        df[yn_target_col] = None
    if pct_target_col not in df.columns:
        df[pct_target_col] = None

    if source_col_name not in df.columns:
        return df

    s = df[source_col_name].astype(str).str.strip()

    def is_yn(val: str) -> bool:
        v = val.lower()
        return v in {"y", "yes", "n", "no", "true", "false", "t", "f", "1", "0"}

    def is_pct(val: str) -> bool:
        m = re.match(r"^\s*(\d+(?:\.\d+)?)\s*%?\s*$", val)
        if not m:
            return False
        try:
            num = float(m.group(1))
            return 0 <= num <= 100
        except Exception:
            return False

    yn_mask = s.apply(is_yn)
    pct_mask = s.apply(is_pct)

    if yn_mask.any():
        df.loc[yn_mask, yn_target_col] = s[yn_mask].str.lower().map({
            "y": "Y", "yes": "Y", "true": "Y", "t": "Y", "1": "Y",
            "n": "N", "no": "N", "false": "N", "f": "N", "0": "N"
        })

    if pct_mask.any():
        pct_vals = s[pct_mask].str.replace("%", "", regex=False).astype(float).clip(0, 100).round().astype(int)
        df.loc[pct_vals.index, pct_target_col] = pct_vals

    return df

# =========================
# AmRisc targets (subset of SOV-APP)
# =========================
TARGETS_IN_ORDER = [
    "* Bldg No.", "*Property Type", "Location Name", "AddressNum", "*Street Address", "*City",
    "*State Code", "*Zip", "County", "Is Prop within 1000 ft of saltwater", "*# of Bldgs",
    "*ISO Const", "Construction Description (provide further details on construction features)",
    "*# of Stories", "*Orig Year Built", "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "*Year Roof covering last fully replaced", "*Real Property Value ($)", "Personal Property Value ($)",
    "M&E (Complete M&E Tech Summary Sheet)", "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "BI/Rental Income ($)", "*Occupancy Description", "Is Prop Mgd or Owned?",
    "Date Added to Sched.", "*# of Units", "*Square Footage", "% Occupied", "Percent Sprinklered",
    "Sprinklered (Y/N)", "ISO Prot Class", "Flood Zone",
]

# Mapping source aliases -> target labels (we normalize keys for matching)
RAW_COLUMN_MAPPING = {
    # * Bldg No.
    "locno": "* Bldg No.",
    "location #": "* Bldg No.",
    "location no": "* Bldg No.",
    "location id": "* Bldg No.",
    "loc #": "* Bldg No.",
    "loc id": "* Bldg No.",
    "building #": "* Bldg No.",
    "bldg #": "* Bldg No.",
    "bldg no": "* Bldg No.",
    "building number": "* Bldg No.",

    # *Property Type
    "property type": "*Property Type",
    "type of property": "*Property Type",
    "asset type": "*Property Type",
    "building type": "*Property Type",
    "structure type": "*Property Type",

    # Location Name
    "location name": "Location Name",
    "location": "Location Name",
    "site name": "Location Name",
    "property name": "Location Name",
    "facility name": "Location Name",
    "building name": "Location Name",

    # *Street Address
    "address": "*Street Address",
    "address 1": "*Street Address",
    "street": "*Street Address",
    "street address": "*Street Address",
    "street name": "*Street Address",
    "location address": "*Street Address",
    "location / address": "*Street Address",
    "st address": "*Street Address",

    # *City
    "city": "*City",
    "city ": "*City",
    "town": "*City",

    # *State Code
    "state": "*State Code",
    "state code": "*State Code",
    "st": "*State Code",
    "state/prov": "*State Code",
    "state/province": "*State Code",
    "province": "*State Code",

    # *Zip
    "zip": "*Zip",
    "zip code": "*Zip",
    "zipcode": "*Zip",
    "postal code": "*Zip",
    "postal": "*Zip",
    "zip code / postal code": "*Zip",

    # County
    "county": "County",
    "county name": "County",
    "parish": "County",

    # Is Prop within 1000 ft of saltwater
    "is prop within 1000 ft of saltwater": "Is Prop within 1000 ft of saltwater",
    "within 1000 ft of saltwater": "Is Prop within 1000 ft of saltwater",
    "saltwater within 1000 ft": "Is Prop within 1000 ft of saltwater",
    "saltwater proximity": "Is Prop within 1000 ft of saltwater",
    "coastal proximity": "Is Prop within 1000 ft of saltwater",
    "within 1000ft saltwater": "Is Prop within 1000 ft of saltwater",

    # *# of Bldgs
    "# buildings": "*# of Bldgs",
    "number of buildings": "*# of Bldgs",
    "num buildings": "*# of Bldgs",
    "ofbuildings": "*# of Bldgs",
    "# of bldgs": "*# of Bldgs",
    "# bldgs": "*# of Bldgs",

    # *ISO Const (type)
    "iso construction": "*ISO Const",
    "iso construction type": "*ISO Const",
    "construction type": "*ISO Const",
    "const type": "*ISO Const",
    "constr type": "*ISO Const",
    "constr. type": "*ISO Const",

    # Construction Description (details)
    "construction": "Construction Description (provide further details on construction features)",
    "construction description": "Construction Description (provide further details on construction features)",
    "air const description": "Construction Description (provide further details on construction features)",
    "const description": "Construction Description (provide further details on construction features)",
    "building construction": "Construction Description (provide further details on construction features)",

    # *# of Stories
    "# stories": "*# of Stories",
    "# of stories": "*# of Stories",
    "number of stories": "*# of Stories",
    "num stories": "*# of Stories",
    "stories": "*# of Stories",
    "ofstories": "*# of Stories",

    # *Orig Year Built
    "year built": "*Orig Year Built",
    "yearbuilt": "*Orig Year Built",
    "yr built": "*Orig Year Built",
    "year": "*Orig Year Built",
    "orig year built": "*Orig Year Built",
    "original year built": "*Orig Year Built",

    # Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)
    "year building upgraded": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "year bldg upgraded": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "major exterior update year": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "renovation year": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "remodel year": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",
    "year remodeled": "Yr Bldg upgraded - Major Exterior Update (mandatory if >25 yrs old)",

    # *Year Roof covering last fully replaced
    "year roof replaced": "*Year Roof covering last fully replaced",
    "roofing year": "*Year Roof covering last fully replaced",
    "roof update year": "*Year Roof covering last fully replaced",
    "roofing update": "*Year Roof covering last fully replaced",
    "roof year": "*Year Roof covering last fully replaced",
    "roof": "*Year Roof covering last fully replaced",

    # *Real Property Value ($)
    "building": "*Real Property Value ($)",
    "buildings": "*Real Property Value ($)",
    "bldg.": "*Real Property Value ($)",
    "bldgs": "*Real Property Value ($)",
    "bldg value": "*Real Property Value ($)",
    "building(s)": "*Real Property Value ($)",
    "real property value ($)": "*Real Property Value ($)",
    "*real property value ($)": "*Real Property Value ($)",
    "building limit": "*Real Property Value ($)",
    "building value": "*Real Property Value ($)",
    "building values": "*Real Property Value ($)",
    "building value ($)": "*Real Property Value ($)",
    "real property": "*Real Property Value ($)",
    "building replacement cost": "*Real Property Value ($)",
    "real property building": "*Real Property Value ($)",
    "building value 26-27": "*Real Property Value ($)",
    "building 26-27": "*Real Property Value ($)",

    # Personal Property Value ($)
    "contents": "Personal Property Value ($)",
    "building content value": "Personal Property Value ($)",
    "contents value": "Personal Property Value ($)",
    "bpp": "Personal Property Value ($)",
    "business personal property limit": "Personal Property Value ($)",
    "business personal property value": "Personal Property Value ($)",
    "business personal property": "Personal Property Value ($)",
    "personal property": "Personal Property Value ($)",
    "contents w/ stock": "Personal Property Value ($)",
    "tib/business personal property limit": "Personal Property Value ($)",
    "business personal property (bpp)" : "Personal Property Value ($)",

    # M&E (Complete M&E Tech Summary Sheet)
    "machinery & equip.": "M&E (Complete M&E Tech Summary Sheet)",
    "machinery and equipment": "M&E (Complete M&E Tech Summary Sheet)",
    "machinery/equipment": "M&E (Complete M&E Tech Summary Sheet)",
    "machinery": "M&E (Complete M&E Tech Summary Sheet)",
    "equipment": "M&E (Complete M&E Tech Summary Sheet)",
    "contractors equipment": "M&E (Complete M&E Tech Summary Sheet)",

    # Other Value $ (outdoor prop & Eqpt must be sch'd)
    "other": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "other values": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "other value": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "container(s)": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "edp": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "electronic data processing": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "miscellaneous": "Other Value $ (outdoor prop & Eqpt must be sch'd)",
    "inventory": "Other Value $ (outdoor prop & Eqpt must be sch'd)",

    # BI/Rental Income ($)
    "bi/ee": "BI/Rental Income ($)",
    "bi/ee value": "BI/Rental Income ($)",
    "bi": "BI/Rental Income ($)",
    "bi ee": "BI/Rental Income ($)",
    "business income limit": "BI/Rental Income ($)",
    "business income w extra expense": "BI/Rental Income ($)",
    "business income/ee": "BI/Rental Income ($)",
    "business income/extra expense": "BI/Rental Income ($)",
    "business interruption & extra expense": "BI/Rental Income ($)",
    "business income/rental income": "BI/Rental Income ($)",
    "bi & ee": "BI/Rental Income ($)",
    "bi w ee": "BI/Rental Income ($)",
    "bi w/ee": "BI/Rental Income ($)",
    "extra expense": "BI/Rental Income ($)",
    "business interruption": "BI/Rental Income ($)",
    "business income": "BI/Rental Income ($)",
    "rents / business income": "BI/Rental Income ($)",
    "rental income": "BI/Rental Income ($)",
    "rents": "BI/Rental Income ($)",
    "rents income & extra exp.": "BI/Rental Income ($)",
    "business income / rents": "BI/Rental Income ($)",
    "effective gross income": "BI/Rental Income ($)",
    "business income (bi), extra expense (ee)" : "BI/Rental Income ($)",

    # *Occupancy Description
    "occupancy": "*Occupancy Description",
    "occupancy description": "*Occupancy Description",
    "occupancy type": "*Occupancy Description",
    "type of occupancy": "*Occupancy Description",
    "building use": "*Occupancy Description",
    "building description": "*Occupancy Description",
    "building type (use)": "*Occupancy Description",
    "air occupancy description": "*Occupancy Description",
    "occupancy - (i.e mixed use, apartments, apartments w/ retail)": "*Occupancy Description",

    # Is Prop Mgd or Owned?
    "is prop mgd or owned?": "Is Prop Mgd or Owned?",
    "managed or owned": "Is Prop Mgd or Owned?",
    "is property managed or owned": "Is Prop Mgd or Owned?",
    "ownership type": "Is Prop Mgd or Owned?",
    "management type": "Is Prop Mgd or Owned?",
    "owned or managed": "Is Prop Mgd or Owned?",

    # Date Added to Sched.
    "date added": "Date Added to Sched.",
    "date added to sched.": "Date Added to Sched.",
    "date added to schedule": "Date Added to Sched.",
    "date added to sov": "Date Added to Sched.",
    "added date": "Date Added to Sched.",

    # *# of Units
    "# units": "*# of Units",
    "# of units": "*# of Units",
    "number of units": "*# of Units",
    "num units": "*# of Units",
    "units": "*# of Units",
    "# of units / containers": "*# of Units",

    # *Square Footage
    "square feet": "*Square Footage",
    "total building square footage": "*Square Footage",
    "total building sf": "*Square Footage",
    "sq ft": "*Square Footage",
    "sq. ft.": "*Square Footage",
    "sqft": "*Square Footage",
    "sf":"*Square Footage",
    "square ft":"*Square Footage",
    "sq. ft":"*Square Footage",
    "building square footage": "*Square Footage",
    "total square footage": "*Square Footage",
    "total sq ft": "*Square Footage",
    "building sqft": "*Square Footage",
    "square footage": "*Square Footage",
    "gross area": "*Square Footage",
    "total area": "*Square Footage",

    # % Occupied
    "% occupied": "% Occupied",
    "occupancy %": "% Occupied",
    "% occupancy": "% Occupied",
    "occupancy percent": "% Occupied",
    "occupancy rate": "% Occupied",

    # Percent Sprinklered
    "percent sprinklered": "Percent Sprinklered",
    "% sprinklered": "Percent Sprinklered",
    "% of structure sprinklered": "Percent Sprinklered",
    "sprinkler (% )": "Percent Sprinklered",
    "% sprinkler coverage": "Percent Sprinklered",

    # Sprinklered (Y/N)
    "sprinklered (y/n)": "Sprinklered (Y/N)",
    "sprinklered?": "Sprinklered (Y/N)",
    "sprinklers (y/n)": "Sprinklered (Y/N)",
    "sprinkler system?": "Sprinklered (Y/N)",
    "sprink y/n/p": "Sprinklered (Y/N)",

    # ISO Prot Class
    "iso protclass": "ISO Prot Class",
    "iso prot class": "ISO Prot Class",
    "protection class": "ISO Prot Class",
    "prot class": "ISO Prot Class",
    "pc": "ISO Prot Class",
    "protectionclass": "ISO Prot Class",

    # Flood Zone
    "flood zone": "Flood Zone",
    "fema flood zone": "Flood Zone",
    "flood": "Flood Zone",
    "flood class": "Flood Zone",
}



# =========================
# Sidebar / Inputs
# =========================
with st.sidebar:
    st.header("⚙️ Settings")
    named_insured = st.text_input(
        "Named Insured (used to name the output file)",
        placeholder="Enter Named Insured...",
    )

    source_sov = st.file_uploader(
        "Upload Source SOV (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="This is the client-provided SOV you want to normalize."
    )

    template_source_choice = st.radio(
        "Template source",
        options=["Use a local/network path"],
        index=0,
    )

    uploaded_template = None
    template_path = None
    if template_source_choice == "Upload template file":
        uploaded_template = st.file_uploader(
            "Upload AmRisc Template (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=False,
            help="If not provided, you can switch to the 'path' option."
        )
    else:
        template_path = st.text_input(
            "Template path",
            value=r"AmRisc_SOV_Schedule.xlsx",
            help="Local or network path accessible from where Streamlit is running."
        )

    template_sheet_name = st.text_input(
        "Template sheet name",
        value="SOV-APP",
        help="The sheet where rows will be written."
    )

    append_to_first_empty = st.checkbox(
        "Append at first empty row under '*Street Address'",
        value=True
    )

    default_start_row = st.number_input(
        "If not appending, start writing at row",
        min_value=1,
        value=3,
        step=1
    )

    save_to_disk = st.checkbox(
        "Also save output to disk (local/network path)", value=False
    )

    output_disk_path = None
    if save_to_disk:
        default_out = rf"U:\Amrisc SOVs\{(named_insured or 'Named Insured').strip() or 'Named Insured'} - Amrisc SOV.xlsx"
        output_disk_path = st.text_input(
            "Output path",
            value=default_out,
            help="Save a copy to this path, if the process has write permissions."
        )

# =========================
# Main UI
# =========================
st.subheader("How it works")
st.markdown(
    """
1. This will detect the column headers in the uploaded Source SOV.
    - Best match results to format column headers: *Building Value, BPP, BI/EE, Square Feet, Occupancy Description, Contruction Type*
2. Type in the Named Insured to name the outputted file
3. Drop or upload your source file that you want to convert into an CrossCover SOV
4. Select "Use a local/network path" this is the Amrisc template
5. Click Process SOV
6. This Transfers columns to the CrossCover template fields.
7. Download the finished Excel.

"""
)

process_button = st.button("🚀 Process SOV", type="primary", use_container_width=True)

# =========================
# Core processing on click
# =========================
if process_button:
    if not source_sov:
        st.error("Please upload a **Source SOV (.xlsx)**.")
        st.stop()
    if template_source_choice == "Upload template file" and not uploaded_template:
        st.error("Please upload a **Template (.xlsx)** or switch to the path option.")
        st.stop()
    if template_source_choice == "Use a local/network path" and not template_path:
        st.error("Please provide a **Template path**.")
        st.stop()
    if not named_insured:
        st.warning("No Named Insured provided. The download name will be generic.")

    try:
        with st.spinner("Reading source SOV and detecting sheet/header…"):
            source_bytes = source_sov.read()
            src_buf_for_detection = io.BytesIO(source_bytes)
            sheet_detected, header_row_index = find_sheet_and_header(src_buf_for_detection)
            st.success(f"Detected sheet: **{sheet_detected}**  header row index (0-based): **{header_row_index}**")

            # Read dataframe using detected sheet/header
            src_df = pd.read_excel(
                io.BytesIO(source_bytes),
                sheet_name=sheet_detected,
                header=header_row_index,
                engine="openpyxl"
            )

            # Try splitting 'City, State, Zip' column by name or pattern
            combined_col_name = None
            for candidate in src_df.columns:
                name_l = str(candidate).strip().lower()
                if name_l in ("city, state, zip", "city,state,zip"):
                    combined_col_name = candidate
                    break
            if combined_col_name is None:
                pattern = re.compile(r"^[^,]+,\s*[A-Za-z]{2},\s*\d{5}(?:-\d{4})?$")
                for candidate in src_df.columns:
                    name_l = str(candidate).strip().lower()
                    if name_l in ("city", "state", "state/prov", "state/province", "zip", "postal code", "postalcode", "zipcode"):
                        continue
                    sample = src_df[candidate].dropna().astype(str).head(50).str.strip()
                    if not sample.empty:
                        m = sample.apply(lambda x: bool(pattern.match(x)))
                        if m.mean() >= 0.8:
                            combined_col_name = candidate
                            break
            if combined_col_name is not None:
                parts = split_city_state_zip_col(src_df[combined_col_name])
                for col in ("City", "State", "Zip"):
                    if col in src_df.columns:
                        mask = is_blank_series(src_df[col]) & parts[col].notna()
                        src_df.loc[mask, col] = parts.loc[mask, col]
                    else:
                        src_df[col] = parts[col]
                st.info(f"Split combined column **'{combined_col_name}'** → City / State / Zip")
            else:
                st.info("No combined 'City, State, Zip' column detected.")

            # Prefer rows with a street-like column populated
            street_candidates = [c for c in src_df.columns if normalize_alias(c) in {"street", "streetaddress", "address"}]
            if street_candidates:
                street_col = street_candidates[0]
                src_df = src_df[src_df[street_col].astype(str).str.strip().ne("")]
            st.write("**Source Data Columns:**", list(src_df.columns))

            # Build new_data with target columns kept even when missing
            new_data = pd.DataFrame(columns=TARGETS_IN_ORDER)

            # Normalize BOTH mapping keys and source columns for robust matching
            src_cols_norm = {normalize_alias(c): c for c in src_df.columns}
            for src_key, tgt_label in RAW_COLUMN_MAPPING.items():
                key_norm = normalize_alias(src_key)
                if key_norm in src_cols_norm and tgt_label in new_data.columns:
                    new_data[tgt_label] = src_df[src_cols_norm[key_norm]].values
            
            # Derive AddressNum from Street Address if not present
            if ("AddressNum" not in new_data.columns) or new_data["AddressNum"].isna().all():
                def addrnum(s):
                    if not isinstance(s, str):
                        return None
                    m = re.match(r"\s*(\d{1,6})\b", s.strip())
                    return m.group(1) if m else None
                street_series = new_data.get("*Street Address")
                if street_series is not None:
                    new_data["AddressNum"] = street_series.apply(addrnum)

            # Normalize State Code (2-letter uppercase)
            if "*State Code" in new_data.columns:
                new_data["*State Code"] = new_data["*State Code"].astype(str).str.upper().str.strip().str[:2]
            
          

            # Ensure Zip is string
            if "*Zip" in new_data.columns:
                new_data["*Zip"] = new_data["*Zip"].astype(str).str.strip()
                        
            # --- ZIP normalizer: ensure clean text ZIP (preserve leading zeros) ---
            if "*Zip" in new_data.columns:
                # Start from text
                z = new_data["*Zip"].astype(str).str.strip()
            
                # Remove spaces and common non-digit noise (including trailing ".0")
                # Keep only digits; this safely handles "02481.0", "02481 ", "02481-1234", etc.
                z_digits = z.str.extract(r'(\d{5}(?:\d{4})?)', expand=False)
            
                # Format as 5-digit or ZIP+4 with hyphen
                def fmt_zip(s):
                    if not isinstance(s, str):
                        return None
                    s = s.strip()
                    if len(s) == 5 and s.isdigit():
                        return s.zfill(5)          # preserve leading zeros
                    if len(s) == 9 and s.isdigit():
                        return f"{s[:5]}-{s[5:]}"  # ZIP+4 as 12345-6789
                    return s if s else None        # leave other cases as-is
            
                new_data["*Zip"] = z_digits.apply(fmt_zip)

            # --- Strict sprinkler mapping: map what's present (no derivation) ---
            sprinkler_source_col = find_sprinkler_col(src_df)
            if sprinkler_source_col:
                temp = src_df[[sprinkler_source_col]].copy()
                temp = map_sprinkler_to_targets(
                    temp,
                    sprinkler_source_col,
                    yn_target_col="Sprinklered (Y/N)",
                    pct_target_col="Percent Sprinklered"
                )
                # Ensure targets exist in new_data
                for col in ["Sprinklered (Y/N)", "Percent Sprinklered"]:
                    if col not in new_data.columns:
                        new_data[col] = None
                # Assign row-by-row
                new_data.loc[:, "Sprinklered (Y/N)"] = temp["Sprinklered (Y/N)"].values
                new_data.loc[:, "Percent Sprinklered"] = temp["Percent Sprinklered"].values
            else:
                for col in ["Sprinklered (Y/N)", "Percent Sprinklered"]:
                    if col not in new_data.columns:
                        new_data[col] = None

            st.markdown("**First 5 rows of mapped data:**")
            st.dataframe(new_data.head(), use_container_width=True)

        # Load template & resolve headers
        with st.spinner("Loading template and resolving headers…"):
            if template_source_choice == "Upload template file":
                template_bytes = uploaded_template.read()
                wb = load_workbook(filename=io.BytesIO(template_bytes))
            else:
                wb = load_workbook(filename=template_path)

            if template_sheet_name not in wb.sheetnames:
                st.error(f"Template sheet **'{template_sheet_name}'** not found. Sheets available: {wb.sheetnames}")
                st.stop()

            ws = wb[template_sheet_name]

            # Detect the real header row (scan top 100 rows)
            targets_norm_set = {normalize_alias(t) for t in TARGETS_IN_ORDER}
            best_row = detect_template_header_row(ws, targets_norm_set, scan_top=100)

            # Build alias -> column index from detected header row
            alias_to_colidx, raw_headers = build_alias_to_colidx(ws, best_row)

            with st.expander("Template header match report", expanded=False):
                st.write(f"Detected header row: {best_row}")
                st.write(raw_headers)
                rows = []
                unmatched = []
                for tgt_label in TARGETS_IN_ORDER:
                    k = normalize_alias(tgt_label)
                    col_idx = alias_to_colidx.get(k)
                    rows.append((tgt_label, k, col_idx if col_idx else "NOT FOUND"))
                    if not col_idx:
                        unmatched.append(tgt_label)
                st.table(pd.DataFrame(rows, columns=["Target Label", "Normalized", "Col Index"]))
                if unmatched:
                    st.warning("Targets not found in template header: " + ", ".join(unmatched))

            # Decide where to write (row selection) - ALWAYS below the detected header row
            baseline_start = best_row + 1
            if append_to_first_empty:
                street_key = normalize_alias("*Street Address")
                street_col = alias_to_colidx.get(street_key)
                start_row = first_empty_row_under(ws, street_col, start=int(baseline_start)) if street_col else int(baseline_start)
            else:
                start_row = max(int(default_start_row), int(baseline_start))
            st.info(f"Writing will start at row: **{start_row}**")

            # Write data locked to resolved columns
            def row_should_be_written(row) -> bool:
                return non_empty(str(row.get("*Street Address", ""))) or (
                    non_empty(str(row.get("*City", ""))) and non_empty(str(row.get("*State Code", "")))
                )

            written = 0
            skipped = 0
            for _, r in new_data.iterrows():
                if not row_should_be_written(r):
                    skipped += 1
                    continue
                target_row = start_row + written
                for tgt_label in TARGETS_IN_ORDER:
                    col_idx = alias_to_colidx.get(normalize_alias(tgt_label))
                    if col_idx:
                        safe_write(ws, target_row, col_idx, r.get(tgt_label))
                written += 1

            # Save to BytesIO for download
            with io.BytesIO() as out_buf:
                wb.save(out_buf)
                out_buf.seek(0)
                safe_name = (named_insured or "Named Insured").strip() or "Named Insured"
                download_name = f"{safe_name} - Amrisc SOV.xlsx"
                st.success(f"Transfer complete ✅  (Rows written: {written}, skipped: {skipped})")
                st.download_button(
                    label="⬇️ Download Completed AmRisc SOV",
                    data=out_buf.getvalue(),
                    file_name=download_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Optionally save to disk (local/network path)
                if save_to_disk and output_disk_path:
                    try:
                        with open(output_disk_path, "wb") as f:
                            f.write(out_buf.getvalue())
                        st.info(f"Also saved a copy to: `{output_disk_path}`")
                    except Exception as e:
                        st.warning(f"Could not save to disk: {e}")

    except Exception as e:
        st.error(f"Processing failed: {e}")






