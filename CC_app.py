# crosscover_sov_app.py
# Streamlit UI wrapper for CrossCover SOV processing
# Keeps the original detection & mapping logic, adds uploads, options, and a download button.

import io
import re
from typing import Tuple, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================
# Streamlit Page Setup
# =========================
st.set_page_config(page_title="CrossCover SOV Builder", page_icon="📄", layout="wide")
st.title("📄 CrossCover SOV Builder")
st.caption("Wraps the original `CrossCover_Sov_fixed.py` into a point-and-click app.")

# =========================
# Helpers (ported & refined)
# =========================
def norm(s: Optional[str]) -> str:
    """Normalize header text for matching."""
    """Lower, trim, collapse whitespace, '&amp;'->'and', remove *,(), strip non-alphanum."""
    if x is None:
        return ""
    s = str(x).strip().lower().replace("&amp;", "and")
    s = re.sub(r"[\*\(\)]", "", s)      # remove *, (, )
    s = re.sub(r"\s+", " ", s)         # collapse whitespace
    return re.sub(r"[^a-z0-9]", "", s)  # strip non-alphanum

def split_lines(s: Optional[str]):
    """Split a cell with embedded line breaks into individual aliases."""
    if not isinstance(s, str):
        return []
    return [p.strip() for p in re.split(r"[\r\n]+", s) if p and p.strip()]

def is_blank_series(series: pd.Series) -> pd.Series:
    return series.isna() | (series.astype(str).str.strip() == "")

def split_city_state_zip_col(series: pd.Series) -> pd.DataFrame:
    """
    Split 'City, ST, 12345' or 'City, ST, 12345-6789' into City/State/Zip.
    Returns a DF with City/State/Zip columns (string dtype).
    """
    s = series.astype("string").str.strip()
    pat = r"^\s*(?P<City>[^,]+?)\s*,\s*(?P<State>[A-Za-z]{2})\s*,\s*(?P<Zip>\d{5}(?:-\d{4})?)\s*$"
    parts = s.str.extract(pat)
    for col in ("City", "State", "Zip"):
        if col in parts:
            parts[col] = parts[col].astype("string").str.strip()
    if "State" in parts:
        parts["State"] = parts["State"].str.upper()
    return parts

def norm_text(x):
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = s.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", s)

# Synonym groups: require at least 3 groups to match for a header row
GROUPS = [
    {"streetaddress", "street", "address", "street+id", "streetid", "streetaddress1"},
    {"city", "town"},
    {"state", "stateprov", "stateprovince", "province"},
    {"zip", "zipcode", "postal", "postalcode"},
]

def looks_like_header(row_vals, min_groups=3) -> bool:
    toks = {norm_text(v) for v in row_vals if isinstance(v, (str, int, float)) and str(v).strip()}
    score = sum(any(alias in toks for alias in group) for group in GROUPS)
    return score >= min_groups

def find_sheet_and_header(xlsx_file_like, search_rows=40) -> Tuple[str, int]:
    """
    Return (sheet_name, header_row_index) by scanning the first `search_rows`
    rows of each sheet for likely address headers.
    """
    # Important: create a fresh BytesIO for each read, since engines may read/seek.
    data = xlsx_file_like.getvalue() if isinstance(xlsx_file_like, io.BytesIO) else xlsx_file_like.read()
    xlsx1 = io.BytesIO(data)
    xlsx2 = io.BytesIO(data)

    xfile = pd.ExcelFile(xlsx1, engine="openpyxl")
    for sheet in xfile.sheet_names:
        raw = pd.read_excel(xlsx2, sheet_name=sheet, header=None, engine="openpyxl")
        for i in range(min(search_rows, len(raw))):
            row_vals = raw.iloc[i].tolist()
            if looks_like_header(row_vals):
                return sheet, i
        xlsx2.seek(0)  # reset for the next sheet
    raise RuntimeError("Could not find a header row in any sheet (looking for Street/City/State/Zip synonyms).")

def first_empty_row_under(ws, column_index: int, start: int = 3) -> int:
    r = start
    while True:
        if ws.cell(row=r, column=column_index).value in (None, ""):
            return r
        r += 1

# ==========================================================
# Target: CrossCover headers we will write (subset of A:Y)
# NOTE: 'Business Interuption' spelling matches the template cell exactly.
# ==========================================================
TARGETS_IN_ORDER = [
    "Location Name", "Street Address", "City", "State", "Zip",
    "Building", "Contents", "Business Interuption", "Machinery & Equip.", "Other",
    "Building SQFT", "Num Buildings", "Num Units", "Num Stories",
    "% Sprinklered", "% Occupied", "Construction", "Construction Description",
    "Occupancy", "Occupancy Description", "ISO ProtClass", "YearBuilt",
]

# ==========================================================
# Column mapping (SOURCE -> TARGET)
# (Same content as your original, with minor whitespace tidy)
# ==========================================================
column_mapping = {
    # Address block
    "Address": "Street Address",
    "ADDRESS": "Street Address",
    "Street": "Street Address",
    "Street Name": "Street Address",
    "*Street Address": "Street Address",
    "Street Address": "Street Address",
    "STREET ADDRESS": "Street Address",
    "Location Address": "Street Address",
    "LOCATION / ADDRESS": "Street Address",
    "streetaddress": "Street Address",
    "StreetAddress": "Street Address",

    "City": "City",
    "CITY": "City",
    "Town": "City",
    "*City": "City",
    "city": "City",

    "State": "State",
    "STATE": "State",
    "ST": "State",
    "St": "State",
    "State/Prov": "State",
    "State/Province": "State",
    "Province": "State",
    "State Code": "State",
    "state": "State",

    "Zip": "Zip",
    "*Zip": "Zip",
    "ZIP": "Zip",
    "ZIP Code": "Zip",
    "Zip code": "Zip",
    "Zip Code": "Zip",
    "Zip Code / Postal Code": "Zip",
    "Postal Code": "Zip",
    "PostalCode": "Zip",
    "Postal": "Zip",
    "zip": "Zip",
    "zipcode": "Zip",
    "ZipCode": "Zip",

    # Values / Exposures
    "Building": "Building",
    " Building ": "Building",
    " Building ": "Building",
    "BUILDING": "Building",
    #"Bldg": "Building",
    "Bldg. Value": "Building",
    "Bldgs": "Building",
    "Bldg Value": "Building",
    "Building(s)": "Building",
    "Buildings [L4]": "Building",
    "Real Property Value ($)": "Building",
    "*Real Property Value ($)": "Building",
    "Building Limit": "Building",
    "Building Value": "Building",
    "Building Values": "Building",
    "Building Value ($)": "Building",
    "2025-2026 Building Value": "Building",
    "Real Property": "Building",
    "Building Replacement Cost": "Building",
    "Building Value (Replacement Cost Valuation)": "Building",
    "Building Insured Value (2025)": "Building",
    "Total Building Value": "Building",
    "buildingvaluereplacementcostvaluation": "Building",

    "Contents": "Contents",
    "Building Content Value": "Contents",
    "Contents Value": "Contents",
    "BPP": "Contents",
    "Business Personal Property Limit": "Contents",
    "Business Personal Property Limit": "Contents",
    "Business Personal Property Value": "Contents",
    "BUSINESS PERSONAL PROPERTY": "Contents",
    "Personal Property Value ($) ": "Contents",
    "Business Personal Property Value ($)": "Contents",
    "Personal Property": "Contents",
    "Business Personal Property": "Contents",
    "Contents w/ Stock": "Contents",
    "TIB/Business Personal Property Limit": "Contents",
    "Personal Property": "Contents",
    "businesspersonalproperty": "Contents",
    "BPP Limit": "Contents",

    "BI/EE": "Business Interuption",
    "BI/EE Value": "Business Interuption",
    "BI": "Business Interuption",
    "BI EE": "Business Interuption",
    "Business Income Limit": "Business Interuption",
    "BI/Rental Income ($)": "Business Interuption",
    "Business Income w Extra Expense": "Business Interuption",
    "Business Income/EE": "Business Interuption",
    "Business Income/Extra Expense": "Business Interuption",
    "Business Interruption & Extra Expense": "Business Interuption",
    "Business Income Limit": "Business Interuption",
    "Business Income/Rental Income": "Business Interuption",
    "BI & EE": "Business Interuption",
    "BI w EE": "Business Interuption",
    "BI w/EE": "Business Interuption",
    "Extra Expense": "Business Interuption",
    "Business Interruption": "Business Interuption",
    "Business Income": "Business Interuption",
    "Rents / Business Income": "Business Interuption",
    "Rental Income": "Business Interuption",
    "Rents": "Business Interuption",
    "Rents Income &  Extra Exp.": "Business Interuption",
    "Business Income / Rents ": "Business Interuption",
    "Effective Gross Income": "Business Interuption",
    "Annual Rental Income": "Business Interuption",
    "businessincomeextraexpense": "Business Interuption",
    "Bus Income Limit": "Business Interuption",
    "Business Income/Extra Expense Limit": "Business Interuption",
    "Business Income Limit": "Business Interuption",
    "Business Income/Extra Expense": "Business Interuption",

    "Machinery & Equip.": "Machinery & Equip.",
    "Machinery and Equipment": "Machinery & Equip.",
    "Machinery/Equipment": "Machinery & Equip.",
    "Machinery": "Machinery & Equip.",
    "Equipment": "Machinery & Equip.",
    "Mach & Equip [L4]": "Machinery & Equip.",
    "Contractors Equipment": "Machinery & Equip.",
    "Machinery/Equipment": "Machinery & Equip.",

    "Other": "Other",
    "Other Values": "Other",
    "Other Value": "Other",
    "Container(s)": "Other",
    "EDP": "Other",
    "EDP (Electronic Data Processing) eg. Computers/printers": "Other",
    "Electronic Data Processing": "Other",
    "Miscellaneous": "Other",
    "Inventory": "Other",

    "Square Feet": "Building SQFT",
    "Total Building Square Footage": "Building SQFT",
    "Total Building SF": "Building SQFT",
    "Sq Ft": "Building SQFT",
    "Sq. Ft.": "Building SQFT",
    "Square feet": "Building SQFT",
    "SqFt": "Building SQFT",
    "sf": "Building SQFT",
    "SF": "Building SQFT",
    "Building Square Footage": "Building SQFT",
    "Total Square Footage": "Building SQFT",
    "Building SQFT": "Building SQFT",
    "Living Area": "Building SQFT",
    "Area": "Building SQFT",
    "Occupied Square Feet": "Building SQFT",
    "Square Footage": "Building SQFT",
    "*Square Footage": "Building SQFT",
    "Sq. Footage": "Building SQFT",
    "Total Sq Ft": "Building SQFT",
    "TOTAL SQ FT": "Building SQFT",
    "Total Area Sq. Ft.": "Building SQFT",
    "Gross Area": "Building SQFT",
    "Total Area": "Building SQFT",
    "Total SQF": "Building SQFT",
    " Building Sq Ft": "Building SQFT",

    # " $/SQFT " omitted here on purpose unless you map it explicitly

    "Num Buildings": "Num Buildings",
    "# Buildings": "Num Buildings",
    "*# of Bldgs": "Num Buildings",
    " #of Buildings ": "Num Buildings",
    "# of Bldgs": "Num Buildings",
    "# Bldgs": "Num Buildings",
    "Number of Bldgs": "Num Buildings",
    "Number of Buildings": "Num Buildings",

    "Num Units": "Num Units",
    "Number of Units": "Num Units",
    "# Units": "Num Units",
    "# of Units": "Num Units",
    "Units": "Num Units",
    "# of Units / Containers": "Num Units",
    "*# of Units": "Num Units",
    "# of Units/Tenants": "Num Units",
    "Num Stories": "Num Stories",
    "# of Stories": "Num Stories",
    "# Stories": "Num Stories",
    "Stories": "Num Stories",
    "*# of Stories": "Num Stories",
    "#of Stories": "Num Stories",
    "Number of Stories": "Num Stories",
    "No of Stories": "Num Stories",
    "No. Stories": "Num Stories",

    "% Sprinklered": "% Sprinklered",
    "Sprinklered": "% Sprinklered",
    "Sprinkler (Y/N)": "% Sprinklered",
    "% Sprinkler Coverage": "% Sprinklered",
    "Sprnkl": "% Sprinklered",
    "Sprink": "% Sprinklered",
    "Sprinklered?": "% Sprinklered",
    "Sprink Y/N/P": "% Sprinklered",
    "Sprinkler": "% Sprinklered",
    "Sprinkler ": "% Sprinklered",
    "Sprinkler (%)": "% Sprinklered",
    "Sprinklers": "% Sprinklered",
    "SPINKLER INFO (Full/Partial)": "% Sprinklered",
    "Warehouse Sprinklered": "% Sprinklered",
    "Spkld?": "% Sprinklered",
    "Sprinkler System?": "% Sprinklered",
    "% of Structure Sprinklered": "% Sprinklered",
    "Sprinklers (Y/N)": "% Sprinklered",

    "% Occupied": "% Occupied",
    "% Occuppied": "% Occupied",
    "Occupancy %": "% Occupied",
    "% Occupancy": "% Occupied",
    "Occupancy Percent": "% Occupied",
    "OCUPANCY PERCENTAGE": "% Occupied",
    "Occupancy Rate": "% Occupied",

    # Construction / Occupancy
    "ISO Construction Type": "Construction Description",
    "ISO Construction": "Construction Description",
    "Construction Description": "Construction Description",
    "Construction": "Construction Description",
    "Const Type": "Construction Description",
    "Constr Type": "Construction Description",
    "Constr. Type *": "Construction Description",
    "Construction Type": "Construction Description",
    "Construction Type": "Construction Description",
    "Type of Construction": "Construction Description",
    "AIR Const Description": "Construction Description",
    "CONSTRUCTION": "Construction Description",
    "Const. Description": "Construction Description",
    "CONSTRUCTION TYPE (i.e. Frame, Masonry Non Combust, Fire Resistive)": "Construction Description",
    "Const": "Construction Description",
    "CONST": "Construction Description",
    "Building Construction": "Construction Description",
    "construction": "Construction Description",

    "Occupancy": "Occupancy Description",
    "occupancy": "Occupancy Description",
    "Occupancy ""Type": "Occupancy Description",
    "OCCUPANCY - (i.e Mixed Use, Apartments, Apartments w/ retail)": "Occupancy Description",
    "OCCUPANCY": "Occupancy Description",
    "*Occupancy": "Occupancy Description",
    "Occupancy Type": "Occupancy Description",
    "Type of Occupancy": "Occupancy Description",
    "Occupancy Description": "Occupancy Description",
    "Description": "Occupancy Description",
    "Building Use": "Occupancy Description",
    "Building Type": "Occupancy Description",
    "Building Description": "Occupancy Description",
    "AIR Occupancy Description": "Occupancy Description",
    "Type of Property": "Occupancy Description",
   # "Location Name": "Occupancy Description",

    "ISO ProtClass": "ISO ProtClass",
    "Protection Class": "ISO ProtClass",
    "Prot Class": "ISO ProtClass",
    "PC": "ISO ProtClass",

    # Year built variations
    "Year Built": "YearBuilt",
    "Year built": "YearBuilt",
    "YearBuilt": "YearBuilt",
    "Yr Built": "YearBuilt",
    "Year Blt": "YearBuilt",
    "Yr. Built": "YearBuilt",
    "Year": "YearBuilt",
    "Built": "YearBuilt",
    "Orig Year Built": "YearBuilt",
    "*Orig Year Built": "YearBuilt",
    "Original Year Built": "YearBuilt",
    "Year Bldt": "YearBuilt",

    # Year roof replaced
    "Year Roof Replaced": "Year Roof Replaced",
    "Roofing Year": "Year Roof Replaced",
    "Roofing": "Year Roof Replaced",
    "Roof Update Year": "Year Roof Replaced",
    "Roof update year": "Year Roof Replaced",
    "Roofing Update": "Year Roof Replaced",
    "Roof Year": "Year Roof Replaced",
    "Roof": "Year Roof Replaced",
    "Remodel Date": "Year Roof Replaced",
}

# =========================
# Sidebar / Inputs
# =========================
with st.sidebar:
    st.header("⚙️ Settings")

    named_insured = st.text_input(
        "Named Insured (used to name the output file)",
        placeholder="Enter Named Insured...",
    )

    source_sov = st.file_uploader(
        "Upload Source SOV (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        help="This is the client-provided SOV you want to normalize."
    )

    template_source_choice = st.radio(
        "Template source",
        options=["Use a local/network path"],
        index=0,
    )

    uploaded_template = None
    template_path = None
    if template_source_choice == "Upload template file":
        uploaded_template = st.file_uploader(
            "Upload CrossCover Template (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=False,
            help="If not provided, you can switch to the 'path' option."
        )
    else:
        template_path = st.text_input(
            "Template path",
            value="CrossCover SOV - Blank.xlsx",
            help="Local or network path accessible from where Streamlit is running."
        )

    template_sheet_name = st.text_input(
        "Template sheet name",
        value="III. New Locations",
        help="The sheet where rows will be written."
    )

    append_to_first_empty = st.checkbox(
        "Append at first empty row under 'Street Address'",
        value=True
    )

    default_start_row = st.number_input(
        "If not appending, start writing at row",
        min_value=1,
        value=3,
        step=1
    )

    save_to_disk = st.checkbox(
        "Also save output to disk (local/network path)", value=False
    )
    output_disk_path = None
    if save_to_disk:
        default_out = r"U:\CrossCover SOVs\{ni} - CrossCover SOV.xlsx".format(
            ni=(named_insured or "Named Insured")
        )
        output_disk_path = st.text_input(
            "Output path",
            value=default_out,
            help="Save a copy to this path, if the process has write permissions."
        )

# =========================
# Main UI
# =========================
st.subheader("How it works")
st.markdown(
    """
1. This will detect the column headers in the uploaded Source SOV.
    - Best match results to format column headers: **Building Value, BPP, BI/EE, Square Feet, Occupancy Description, Contruction Type**
2. Type in the Named Insured to name the outputted file
3. Drop or upload your source file that you want to convert into an CrossCover SOV
4. Select **"Use a local/netowrk path"** this is the CrossCover template
5. Click **Process SOV**
6. This **Transfers** columns to the CrossCover template fields.    
7. Download the **finished Excel**.
"""
)

process_button = st.button("🚀 Process SOV", type="primary", use_container_width=True)

# =========================
# Core processing on click
# =========================
if process_button:
    if not source_sov:
        st.error("Please upload a **Source SOV (.xlsx)**.")
        st.stop() 
    if template_source_choice == "Use a local/network path" and not template_path:
        st.error("Please provide a **Template path**.")
        st.stop()
    if template_source_choice == "Upload template file" and not uploaded_template:
        st.error("Please upload a **Template (.xlsx)** or switch to the path option.")
        st.stop()
    if not named_insured:
        st.warning("No Named Insured provided. The download name will be generic.")

    try:
        with st.spinner("Reading source SOV and detecting sheet/header…"):
            # Read source bytes once; reuse safely
            source_bytes = source_sov.read()
            src_buf_for_detection = io.BytesIO(source_bytes)
            sheet_detected, header_row_index = find_sheet_and_header(src_buf_for_detection)
            st.success(f"Detected sheet: **{sheet_detected}** | header row index (0-based): **{header_row_index}**")

            # Now read the dataframe using detected sheet/header
            src_df = pd.read_excel(
                io.BytesIO(source_bytes),
                sheet_name=sheet_detected,
                header=header_row_index,
                engine="openpyxl"
            )

        # Try to split a combined 'City, ST, Zip' column if present
        combined_col_name = None
        for candidate in src_df.columns:
            if str(candidate).strip().lower() in ("city, state, zip", "city,state,zip"):
                combined_col_name = candidate
                break

        if combined_col_name is None:
            # Heuristic scan for a column shaped like "City, ST, Zip"
            pattern = re.compile(r"^[^,]+,\s*[A-Za-z]{2},\s*\d{5}(?:-\d{4})?$")
            for candidate in src_df.columns:
                name_l = str(candidate).strip().lower()
                if name_l in ("city", "state", "state/prov", "state/province", "zip", "postal code", "postalcode", "zipcode"):
                    continue
                sample = src_df[candidate].dropna().astype(str).head(50).str.strip()
                if not sample.empty:
                    m = sample.apply(lambda x: bool(pattern.match(x)))
                    if m.mean() >= 0.8:
                        combined_col_name = candidate
                        break

        if combined_col_name is not None:
            parts = split_city_state_zip_col(src_df[combined_col_name])
            for col in ("City", "State", "Zip"):
                if col in src_df.columns:
                    mask = is_blank_series(src_df[col]) & parts[col].notna()
                    src_df.loc[mask, col] = parts.loc[mask, col]
                else:
                    src_df[col] = parts[col]
            st.info(f"Split combined column **'{combined_col_name}'** → City / State / Zip")
  #      else:
         #   st.info("No combined 'City, State, Zip' column detected.")

        # Prefer rows with a street-like column populated
        street_candidates = [c for c in src_df.columns if norm(c) in {"street", "streetaddress", "address"}]
        if street_candidates:
            street_col = street_candidates[0]
            src_df = src_df[src_df[street_col].astype(str).str.strip().ne("")]
      #  st.write("**Source Data Columns:**", list(src_df.columns))

        # Build new_data with target columns kept even when missing
        new_data = pd.DataFrame(columns=TARGETS_IN_ORDER)
        for src_col, tgt_label in column_mapping.items():
            if src_col in src_df.columns and tgt_label in new_data.columns:
                new_data[tgt_label] = src_df[src_col].values

        st.markdown("**First 5 rows of mapped data:**")
        st.dataframe(new_data.head(), use_container_width=True)

        # Load template workbook
        with st.spinner("Loading template and resolving headers…"):
            if template_source_choice == "Upload template file":
                template_bytes = uploaded_template.read()
                wb = load_workbook(filename=io.BytesIO(template_bytes))
            else:
                wb = load_workbook(filename=template_path)

            if template_sheet_name not in wb.sheetnames:
                st.error(f"Template sheet **'{template_sheet_name}'** not found. Sheets available: {wb.sheetnames}")
                st.stop()

            ws = wb[template_sheet_name]

            # Build alias -> column index from row 2 (handle wrapped headers)
            alias_to_colidx = {}
            raw_headers = [cell.value for cell in ws[2]]
            for idx, header in enumerate(raw_headers, start=1):
                if header is None:
                    continue
                aliases = set()
                aliases.add(header)  # full text
                if isinstance(header, str):
                    aliases.add(header.replace("\n", " ").replace("\r", " "))
                    for part in split_lines(header):
                        aliases.add(part)
                # include '&' variants
                more = set()
                for a in aliases:
                    if isinstance(a, str):
                        more.add(a.replace("&", "and"))
                aliases |= more

                for a in aliases:
                    key = norm(a)
                    if key and key not in alias_to_colidx:
                        alias_to_colidx[key] = idx

            # Report match status
            with st.expander("Template header match report (row 2)", expanded=False):
                st.write(raw_headers)
                rows = []
                unmatched = []
                for tgt_label in TARGETS_IN_ORDER:
                    k = norm(tgt_label)
                    col_idx = alias_to_colidx.get(k)
                    rows.append((tgt_label, k, col_idx if col_idx else "NOT FOUND"))
                    if not col_idx:
                        unmatched.append(tgt_label)
                st.table(pd.DataFrame(rows, columns=["Target Label", "Normalized", "Col Index"]))
                if unmatched:
                    st.warning("Targets not found in template row 2: " + ", ".join(unmatched))

        # Decide where to write (row selection)
        if append_to_first_empty:
            street_key = norm("Street Address")
            street_col = alias_to_colidx.get(street_key)
            start_row = first_empty_row_under(ws, street_col, start=int(default_start_row)) if street_col else int(default_start_row)
        else:
            start_row = int(default_start_row)

        st.info(f"Writing will start at row: **{start_row}**")

        # Write data locked to resolved columns
        for r_idx, row in new_data.iterrows():
            target_row = start_row + r_idx
            for tgt_label in TARGETS_IN_ORDER:
                k = norm(tgt_label)
                col_idx = alias_to_colidx.get(k)
                if col_idx:
                    ws.cell(row=target_row, column=col_idx, value=row.get(tgt_label))

        # Save to BytesIO for download
        with io.BytesIO() as out_buf:
            wb.save(out_buf)
            out_buf.seek(0)
            safe_name = (named_insured or "Named Insured").strip() or "Named Insured"
            download_name = f"{safe_name} - CrossCover SOV.xlsx"

            st.success("Transfer complete ✅")
            st.download_button(
                label="⬇️ Download Completed CrossCover SOV",
                data=out_buf.getvalue(),
                file_name=download_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # Optionally save to disk (local/network path)
            if save_to_disk and output_disk_path:
                try:
                    with open(output_disk_path, "wb") as f:
                        f.write(out_buf.getvalue())
                    st.info(f"Also saved a copy to: `{output_disk_path}`")
                except Exception as e:
                    st.warning(f"Could not save to disk: {e}")

    except Exception as e:
        st.error(f"Processing failed: {e}")
        st.exception(e)







